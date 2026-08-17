"""xlsx の**決定的**な数値集計（LLM に数を数えさせない）。

集計を LLM に任せると数値のねつ造が必ず起きる。ここで openpyxl から列ごとの
件数・合計・最小・最大・平均を **Python で** 計算し、LLM には「その表を日本語で
整えるだけ」をやらせる（設計監査 fix #7 の 2 段構成）。

read_only=True / data_only=True でストリーミング読み。行・列・シート数に hard cap を
置き、巨大ブックでも有界時間で終わる。
"""

from __future__ import annotations

from dataclasses import dataclass, field
from io import BytesIO

# 有界化（office_extract の上限群と同じ思想）。
MAX_AGG_SHEETS = 10
MAX_AGG_ROWS = 5_000
MAX_AGG_COLUMNS = 30


@dataclass(frozen=True)
class ColumnStats:
    """数値列 1 本の決定的集計。"""

    header: str
    count: int
    total: float
    minimum: float
    maximum: float

    @property
    def mean(self) -> float:
        return self.total / self.count if self.count else 0.0


@dataclass
class SheetStats:
    """シート 1 枚の決定的集計。"""

    title: str
    rows_scanned: int = 0
    truncated: bool = False
    columns: list[ColumnStats] = field(default_factory=list)


def _fmt(value: float) -> str:
    """整数なら整数で、そうでなければ小数第2位までで表示（決定的）。"""
    if value == int(value) and abs(value) < 1e15:
        return f"{int(value):,}"
    return f"{value:,.2f}"


def compute_xlsx_stats(
    data: bytes,
    *,
    max_sheets: int = MAX_AGG_SHEETS,
    max_rows: int = MAX_AGG_ROWS,
    max_columns: int = MAX_AGG_COLUMNS,
) -> list[SheetStats]:
    """xlsx バイナリから列ごとの数値集計を決定的に計算する。

    1 行目をヘッダ行とみなす（値が無い列は「列N」）。bool は数値に数えない
    （True を 1 と数えて「合計」を作らない）。日付も数値集計から外す。
    """
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    try:
        out: list[SheetStats] = []
        for ws in wb.worksheets[:max_sheets]:
            stats = SheetStats(title=str(ws.title))
            headers: list[str] = []
            counts: list[int] = []
            totals: list[float] = []
            mins: list[float] = []
            maxs: list[float] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx > max_rows:
                    stats.truncated = True
                    break
                cells = list(row)[:max_columns]
                if row_idx == 1:
                    headers = [
                        (str(c).strip() if c is not None and str(c).strip() else f"列{i + 1}")
                        for i, c in enumerate(cells)
                    ]
                    counts = [0] * len(headers)
                    totals = [0.0] * len(headers)
                    mins = [0.0] * len(headers)
                    maxs = [0.0] * len(headers)
                    continue
                stats.rows_scanned += 1
                for i, cell in enumerate(cells):
                    if i >= len(headers):
                        break
                    if isinstance(cell, bool) or not isinstance(cell, (int, float)):
                        continue
                    v = float(cell)
                    if counts[i] == 0:
                        mins[i] = v
                        maxs[i] = v
                    else:
                        mins[i] = min(mins[i], v)
                        maxs[i] = max(maxs[i], v)
                    counts[i] += 1
                    totals[i] += v
            stats.columns = [
                ColumnStats(
                    header=headers[i],
                    count=counts[i],
                    total=totals[i],
                    minimum=mins[i],
                    maximum=maxs[i],
                )
                for i in range(len(headers))
                if counts[i] > 0
            ]
            out.append(stats)
        return out
    finally:
        wb.close()


def format_stats_ja(sheets: list[SheetStats]) -> str:
    """決定的集計を日本語の表テキストへ整形する（この文字列が数値の真実源）。"""
    if not sheets:
        return ""
    lines: list[str] = []
    for s in sheets:
        head = f"■ シート「{s.title}」データ行 {s.rows_scanned} 行"
        if s.truncated:
            head += f"（先頭 {MAX_AGG_ROWS} 行のみ集計）"
        lines.append(head)
        if not s.columns:
            lines.append("  数値列は見つかりませんでした")
            continue
        for c in s.columns:
            lines.append(
                f"  ・{c.header}: 件数 {c.count} / 合計 {_fmt(c.total)} / "
                f"平均 {_fmt(c.mean)} / 最小 {_fmt(c.minimum)} / 最大 {_fmt(c.maximum)}"
            )
    return "\n".join(lines)


__all__ = [
    "MAX_AGG_COLUMNS",
    "MAX_AGG_ROWS",
    "MAX_AGG_SHEETS",
    "ColumnStats",
    "SheetStats",
    "compute_xlsx_stats",
    "format_stats_ja",
]
